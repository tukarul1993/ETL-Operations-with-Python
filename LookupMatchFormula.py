import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('D:\Python\ETL Operations with Python\Files\IfElseExamples.xlsx')

# Select the worksheet
worksheet = workbook['Sheet2']

# Get the maximum number of rows in column A
max_row = worksheet.max_row


for row in range(5, max_row + 1):
    print(row)
    #gender_cell = f'E{row}'
    #salary_cell = f'H{row}'
    lookup_cell=f'B{row}'
    #=IF(AND(E11="Female",H11<50000),"Eligible for Gift","Not Eligible for Gift")
    #print(gender_cell)
    #formula = f'=IF(AND({gender_cell}="Female",{salary_cell}<50000),"Eligible for Gift","Not Eligible for Gift")'  # {lookup_value}
    formula=f'=MATCH("Basic Salary",$N$4:$Q$4,0)'
    formula2=f'=VLOOKUP({lookup_cell},$N$5:$Q$39,MATCH("Basic Salary",$N$4:$Q$4,0),0)'
    worksheet.cell(row=row, column=11).value = formula

workbook.save('D:\Python\ETL Operations with Python\Files\LookupMatchExampleOp.xlsx')