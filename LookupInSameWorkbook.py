import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook('/Files/VlookupSample2.xlsx')

# Select the worksheets we want to work with
sheet1 = workbook['Sheet1']
sheet2 = workbook['Sheet2']
sheet3 = workbook.create_sheet('Sheet3')

# Loop over the rows in Sheet1
for row in sheet1.iter_rows(min_row=2, values_only=True):
    # Extract the values we want to look up
    id = row[0]
    name = row[1]

    # Look up the corresponding value in Sheet2
    for row2 in sheet2.iter_rows(min_row=2, values_only=True):
        if row2[0] == id:
            value = row2[1]
            break
    else:
        value = None

    # Write the results to Sheet3
    sheet3.append((id, name, value))

# Save the updated workbook
workbook.save('D:\Python\ETL\Files\example.xlsx')
