import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('D:\Python\ETL Operations with Python\Files\IfElseExamples.xlsx')

# Select the worksheet
worksheet = workbook['Sheet1']

# Get the maximum number of rows in column A
max_row = worksheet.max_row


for row in range(11, max_row + 1):
    print(row)
    gender_cell = f'E{row}'
    salary_cell = f'H{row}'
    #=IF(AND(E11="Female",H11<50000),"Eligible for Gift","Not Eligible for Gift")
    print(gender_cell)
    formula = f'=IF(AND({gender_cell}="Female",{salary_cell}<50000),"Eligible for Gift","Not Eligible for Gift")'  # {lookup_value}
    worksheet.cell(row=row, column=16).value = formula

workbook.save('D:\Python\ETL Operations with Python\Files\IfElseExamplesOutput.xlsx')