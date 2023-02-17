import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('D:\Python\ETL Operations with Python\Files\PythonExcelLookupExample.xlsx')

# Select the worksheet
worksheet = workbook['Sheet1']

# Get the maximum number of rows in column A
max_row = worksheet.max_row

# Apply the VLOOKUP formula in all cells of column C, using the appropriate lookup values and ranges
for row in range(2, max_row + 1):
    lookup_value = worksheet.cell(row=row, column=1).value
    lookup_value_cell = f'A{row}'
    print(lookup_value_cell)
    formula = f'=VLOOKUP( {lookup_value_cell}, Sheet2!A2:B6, 2, FALSE)' #{lookup_value}
    print(formula)
    worksheet.cell(row=row, column=3).value = formula

# Save the workbook
workbook.save('D:\Python\ETL Operations with Python\Files\PythonExcelLookupExample2.xlsx')